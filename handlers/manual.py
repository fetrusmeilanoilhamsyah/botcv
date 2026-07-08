"""
handlers/manual.py — Fitur pembuatan kontak secara manual (TXT, VCF, EXCEL).
Mendukung input teks manual dari berbagai negara, tanpa emoji, dan dengan UI premium.
"""
import os
import io
import re
import asyncio
import logging
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import ContextTypes
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

from database import db
from database.db_async import adb
from middleware.auth import require_member
from core.vcf_parser import add_plus
from core.utils import sanitize_filename

logger = logging.getLogger(__name__)

# States untuk in-memory session
S_WAIT_TEXT = "MANUAL_WAIT_TEXT"
S_WAIT_FORMAT = "MANUAL_WAIT_FORMAT"
S_WAIT_CONTACTNAME = "MANUAL_WAIT_CONTACTNAME"
S_WAIT_FILENAME = "MANUAL_WAIT_FILENAME"

_user_locks: dict = {}
_user_timers: dict = {}

def get_user_lock(user_id: int) -> asyncio.Lock:
    return _user_locks.setdefault(user_id, asyncio.Lock())

def cleanup_inactive_users(inactive_ids: list) -> int:
    cleaned = 0
    for uid in inactive_ids:
        _user_locks.pop(uid, None)
        timer = _user_timers.pop(uid, None)
        if timer:
            timer.cancel()
        cleaned += 1
    return cleaned

def extract_numbers_from_text(text: str) -> list[str]:
    if not text:
        return []
        
    segments = re.split(r'[\n\r,;|\t]+', text)
    clean_numbers = []
    seen = set()
    
    for seg in segments:
        seg = seg.strip()
        if not seg:
            continue
        
        seg = re.sub(r'^(?:\d+[\.\)\:\s\-]+|[\-\*\+\•]\s*|no\.\s*\d+\s*[\.\:\-]*\s*)', '', seg, flags=re.IGNORECASE).strip()
        if not seg:
            continue
        
        has_plus = seg.startswith("+")
        digits = re.sub(r'\D', '', seg)
        if not digits:
            continue
        
        cleaned = ("+" if has_plus else "") + digits
        
        if 7 <= len(cleaned) <= 17:
            if cleaned not in seen:
                seen.add(cleaned)
                clean_numbers.append(cleaned)
                
    return clean_numbers

def _get_breadcrumbs(data: dict, step: int) -> str:
    fmt = data.get("format", "")
    numbers_count = len(data.get("numbers", []))
    contact_name = data.get("contact_name", "")
    file_name = data.get("file_name", "")

    parts = []
    
    if step == 1:
        parts.append(f"<b>[NOMOR: {numbers_count}]</b>" if numbers_count else "<b>[INPUT NOMOR]</b>")
    else:
        parts.append(f"Nomor: <code>{numbers_count}</code>" if numbers_count else "Nomor: ➖")
        
    if step == 2:
        parts.append(f"<b>[FORMAT: {fmt.upper()}]</b>" if fmt else "<b>[FORMAT OUTPUT]</b>")
    elif step > 2 and fmt:
        parts.append(f"Format: <code>{fmt.upper()}</code>")
    else:
        parts.append("Format: ➖")
        
    if fmt == "vcf":
        if step == 3:
            parts.append(f"<b>[NAMA: {contact_name.upper()}]</b>" if contact_name else "<b>[NAMA KONTAK]</b>")
        elif step > 3 and contact_name:
            parts.append(f"Nama: <code>{contact_name}</code>")
        else:
            parts.append("Nama: ➖")
            
    actual_filename_step = 4 if fmt == "vcf" else 3
    if step == actual_filename_step:
        parts.append(f"<b>[FILE: {file_name.upper()}]</b>" if file_name else "<b>[NAMA FILE]</b>")
    elif step > actual_filename_step and file_name:
        parts.append(f"File: <code>{file_name}</code>")
    else:
        parts.append("File: ➖")

    breadcrumbs = " ➔ ".join(parts)
    
    console_title = (
        "[ MANUAL VCF CONSOLE ]" if fmt == "vcf" else
        "[ MANUAL TXT CONSOLE ]" if fmt == "txt" else
        "[ MANUAL EXCEL CONSOLE ]" if fmt == "excel" else
        "[ MANUAL CONSOLE ]"
    )
    return (
        f"<b>{console_title}</b>\n"
        "────────────────────────────\n"
        f"<blockquote>{breadcrumbs}</blockquote>\n"
        "────────────────────────────\n\n"
    )

def _waiting_text(data: dict) -> str:
    return (
        _get_breadcrumbs(data, 1) +
        f"<blockquote><b>[ STATUS: WAITING FOR TEXT ]</b>\n"
        f"Silakan ketik atau tempel daftar nomor HP Anda di chat sekarang.\n\n"
        f"<b>Format:</b>\n"
        f"\u2022 Satu nomor per baris, ATAU\n"
        f"\u2022 Dipisahkan dengan koma/spasi/tab</blockquote>"
    )

async def cmd_manual(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await require_member(update, context):
        return
    user_id = update.effective_user.id
    asyncio.create_task(adb.increment_usage(user_id))
    
    from handlers.start import transition_to_handler
    
    init_data = {"numbers": []}
    db.set_session(user_id, S_WAIT_TEXT, init_data)
    
    msg = await transition_to_handler(
        context.bot,
        user_id,
        update.effective_chat.id,
        _waiting_text(init_data),
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("BATAL & KEMBALI", callback_data="back_to_start", style="danger")]]),
        update=update
    )
    if msg:
        sess = db.get_session(user_id)
        sess["data"]["status_msg_id"] = msg.message_id
        db.set_session(user_id, S_WAIT_TEXT, sess["data"])

async def handle_manual_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    sess = db.get_session(user_id)
    if not sess or sess.get("state") != S_WAIT_TEXT:
        return

    text = update.message.text
    try:
        await update.message.delete()
    except Exception:
        pass

    data = sess["data"]
    status_msg_id = data.get("status_msg_id")

    if not text or not text.strip():
        # Input kosong — in-place warning
        try:
            if status_msg_id:
                await context.bot.edit_message_text(
                    chat_id=update.effective_chat.id,
                    message_id=status_msg_id,
                    text=(
                        _get_breadcrumbs(data, 1) +
                        "<blockquote>⚠️ <b>[ KOSONG ]</b>\nTeks yang dikirim kosong. Silakan kirimkan kembali daftar nomor HP Anda.</blockquote>"
                    ),
                    parse_mode="HTML",
                    reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("BATAL & KEMBALI", callback_data="back_to_start", style="danger")]])
                )
                await asyncio.sleep(10)
                await context.bot.edit_message_text(
                    chat_id=update.effective_chat.id,
                    message_id=status_msg_id,
                    text=_waiting_text(data),
                    parse_mode="HTML",
                    reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("BATAL & KEMBALI", callback_data="back_to_start", style="danger")]])
                )
        except Exception:
            pass
        return

    clean_numbers = extract_numbers_from_text(text)

    if not clean_numbers:
        # Tidak ada nomor valid — in-place warning
        try:
            if status_msg_id:
                await context.bot.edit_message_text(
                    chat_id=update.effective_chat.id,
                    message_id=status_msg_id,
                    text=(
                        _get_breadcrumbs(data, 1) +
                        "<blockquote>⚠️ <b>[ TIDAK ADA NOMOR VALID ]</b>\nTidak ditemukan nomor HP yang valid pada teks tersebut. Silakan coba lagi.</blockquote>"
                    ),
                    parse_mode="HTML",
                    reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("BATAL & KEMBALI", callback_data="back_to_start", style="danger")]])
                )
                await asyncio.sleep(10)
                await context.bot.edit_message_text(
                    chat_id=update.effective_chat.id,
                    message_id=status_msg_id,
                    text=_waiting_text(data),
                    parse_mode="HTML",
                    reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("BATAL & KEMBALI", callback_data="back_to_start", style="danger")]])
                )
        except Exception:
            pass
        return

    data["numbers"] = clean_numbers
    db.set_session(user_id, S_WAIT_FORMAT, data)

    keyboard = InlineKeyboardMarkup([
        [
            InlineKeyboardButton("TXT", callback_data="manual_fmt_txt", style="primary"),
            InlineKeyboardButton("VCF", callback_data="manual_fmt_vcf", style="primary"),
            InlineKeyboardButton("EXCEL", callback_data="manual_fmt_excel", style="primary")
        ],
        [InlineKeyboardButton("BATAL & KEMBALI", callback_data="back_to_start", style="danger")]
    ])

    try:
        await context.bot.edit_message_text(
            chat_id=update.effective_chat.id,
            message_id=status_msg_id,
            text=_get_breadcrumbs(data, 2) + f"<blockquote><b>[ LANGKAH 2: PILIH FORMAT OUTPUT ]</b>\nDitemukan: <code>{len(clean_numbers)}</code> nomor HP unik.\n\nPilih format output:</blockquote>",
            parse_mode="HTML",
            reply_markup=keyboard
        )
    except Exception:
        pass

async def handle_manual_format_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    sess = db.get_session(user_id)
    if not sess or sess.get("state") != S_WAIT_FORMAT:
        return

    fmt = query.data.split("_")[-1]
    data = sess["data"]
    data["format"] = fmt
    status_msg_id = data.get("status_msg_id")

    keyboard = InlineKeyboardMarkup([[InlineKeyboardButton("BATAL & KEMBALI", callback_data="back_to_start", style="danger")]])

    if fmt in ("txt", "excel"):
        db.set_session(user_id, S_WAIT_FILENAME, data)
        try:
            await query.edit_message_text(
                text=_get_breadcrumbs(data, 3) + f"<blockquote><b>[ LANGKAH 3: INPUT NAMA FILE ]</b>\nMasukkan nama file hasil {fmt.upper()} (contoh: <code>Hasil</code>):</blockquote>",
                parse_mode="HTML",
                reply_markup=keyboard
            )
        except Exception:
            pass
    elif fmt == "vcf":
        db.set_session(user_id, S_WAIT_CONTACTNAME, data)
        try:
            await query.edit_message_text(
                text=_get_breadcrumbs(data, 3) + f"<blockquote><b>[ LANGKAH 3: INPUT NAMA KONTAK ]</b>\nMasukkan nama kontak untuk hasil VCF (contoh: <code>FEE</code>):</blockquote>",
                parse_mode="HTML",
                reply_markup=keyboard
            )
        except Exception:
            pass

async def handle_manual_contact_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    sess = db.get_session(user_id)
    if not sess or sess.get("state") != S_WAIT_CONTACTNAME:
        return

    contact_name = update.message.text.strip()
    try:
        await update.message.delete()
    except Exception:
        pass

    data = sess["data"]
    status_msg_id = data.get("status_msg_id")
    keyboard = InlineKeyboardMarkup([[InlineKeyboardButton("BATAL & KEMBALI", callback_data="back_to_start", style="danger")]])

    if not contact_name:
        try:
            await context.bot.edit_message_text(
                chat_id=update.effective_chat.id,
                message_id=status_msg_id,
                text=_get_breadcrumbs(data, 3) + "<blockquote>⚠️ <b>Masukkan nama kontak yang valid.</b></blockquote>",
                parse_mode="HTML",
                reply_markup=keyboard
            )
        except Exception:
            pass
        return

    data["contact_name"] = contact_name
    db.set_session(user_id, S_WAIT_FILENAME, data)

    try:
        await context.bot.edit_message_text(
            chat_id=update.effective_chat.id,
            message_id=status_msg_id,
            text=_get_breadcrumbs(data, 4) + "<blockquote><b>[ LANGKAH 4: INPUT NAMA FILE ]</b>\nMasukkan nama file untuk hasil VCF (contoh: <code>FEE_Manual</code>):</blockquote>",
            parse_mode="HTML",
            reply_markup=keyboard
        )
    except Exception:
        pass

async def handle_manual_file_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    sess = db.get_session(user_id)
    if not sess or sess.get("state") != S_WAIT_FILENAME:
        return

    filename = sanitize_filename(update.message.text.strip())
    try:
        await update.message.delete()
    except Exception:
        pass

    data = sess["data"]
    status_msg_id = data.get("status_msg_id")
    keyboard = InlineKeyboardMarkup([[InlineKeyboardButton("BATAL & KEMBALI", callback_data="back_to_start", style="danger")]])

    if not filename:
        try:
            await context.bot.edit_message_text(
                chat_id=update.effective_chat.id,
                message_id=status_msg_id,
                text=_get_breadcrumbs(data, 4 if data.get("format") == "vcf" else 3) + "<blockquote>⚠️ <b>Masukkan nama file yang valid.</b></blockquote>",
                parse_mode="HTML",
                reply_markup=keyboard
            )
        except Exception:
            pass
        return

    data["file_name"] = filename
    db.clear_session(user_id)
    await handle_manual_process(update, context, data)

async def handle_manual_process(update: Update, context: ContextTypes.DEFAULT_TYPE, data: dict):
    user_id = update.effective_user.id
    fmt = data.get("format")
    numbers = data.get("numbers", [])
    file_name = data.get("file_name", "kontak")
    status_msg_id = data.get("status_msg_id")

    process_text = "<blockquote><b>[ SYSTEM: PROCESSING DATA ]</b>\nSedang mengonversi dan membuat file...</blockquote>"
    try:
        await context.bot.edit_message_text(
            chat_id=update.effective_chat.id,
            message_id=status_msg_id,
            text=process_text,
            parse_mode="HTML"
        )
    except Exception:
        pass

    try:
        loop = asyncio.get_running_loop()

        if fmt == "txt":
            def do_txt():
                content = ("\n".join(numbers) + "\n").encode("utf-8")
                return content
            
            content = await loop.run_in_executor(None, do_txt)
            buf = io.BytesIO(content)
            buf.name = f"{file_name}.txt"

            # Kirim status mengirim
            try:
                await context.bot.edit_message_text(
                    chat_id=update.effective_chat.id,
                    message_id=status_msg_id,
                    text="<blockquote><b>[ SYSTEM: SENDING FILES ]</b>\nSedang mengirim berkas TXT hasil...</blockquote>",
                    parse_mode="HTML"
                )
            except Exception:
                pass

            await context.bot.send_document(
                chat_id=update.effective_chat.id,
                document=buf,
                filename=f"{file_name}.txt"
            )

        elif fmt == "excel":
            def do_excel():
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Daftar Kontak"

                header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                header_align = Alignment(horizontal="center", vertical="center")

                headers = ["No", "Nama Kontak", "Nomor HP"]
                ws.append(headers)

                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align

                row_align_center = Alignment(horizontal="center", vertical="center")
                row_align_left = Alignment(horizontal="left", vertical="center")

                for idx, num in enumerate(numbers, 1):
                    ws.append([idx, f"Kontak {idx}", num])
                    ws.cell(row=idx+1, column=1).alignment = row_align_center
                    ws.cell(row=idx+1, column=2).alignment = row_align_left
                    ws.cell(row=idx+1, column=3).alignment = row_align_center

                for col in ws.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = openpyxl.utils.get_column_letter(col[0].column)
                    ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)
                return buf.getvalue()

            content = await loop.run_in_executor(None, do_excel)
            buf = io.BytesIO(content)
            buf.name = f"{file_name}.xlsx"

            try:
                await context.bot.edit_message_text(
                    chat_id=update.effective_chat.id,
                    message_id=status_msg_id,
                    text="<blockquote><b>[ SYSTEM: SENDING FILES ]</b>\nSedang mengirim berkas EXCEL hasil...</blockquote>",
                    parse_mode="HTML"
                )
            except Exception:
                pass

            await context.bot.send_document(
                chat_id=update.effective_chat.id,
                document=buf,
                filename=f"{file_name}.xlsx"
            )

        elif fmt == "vcf":
            contact_name = data.get("contact_name", "FEE")

            def do_vcf():
                vcf_lines = []
                for idx, num in enumerate(numbers, 1):
                    name = f"{contact_name}{idx}"
                    vcf_lines.append(f"BEGIN:VCARD\nVERSION:3.0\nFN:{name}\nTEL;TYPE=CELL:{add_plus(num)}\nEND:VCARD")
                
                content = ("\n".join(vcf_lines) + "\n").encode("utf-8")
                return content

            content = await loop.run_in_executor(None, do_vcf)
            buf = io.BytesIO(content)
            buf.name = f"{file_name}.vcf"

            try:
                await context.bot.edit_message_text(
                    chat_id=update.effective_chat.id,
                    message_id=status_msg_id,
                    text="<blockquote><b>[ SYSTEM: SENDING FILES ]</b>\nSedang mengirim berkas VCF hasil...</blockquote>",
                    parse_mode="HTML"
                )
            except Exception:
                pass

            await context.bot.send_document(
                chat_id=update.effective_chat.id,
                document=buf,
                filename=f"{file_name}.vcf"
            )

        try:
            await context.bot.delete_message(chat_id=update.effective_chat.id, message_id=status_msg_id)
        except Exception:
            pass

        keyboard = InlineKeyboardMarkup([
            [
                InlineKeyboardButton("PROSES MANUAL LAIN", callback_data="show_manual_help", style="success"),
                InlineKeyboardButton("KEMBALI KE MENU", callback_data="back_to_start", style="danger")
            ]
        ])

        def _fit(val, max_len=22) -> str:
            s = str(val)
            if len(s) > max_len:
                return s[:max_len-3] + "..."
            return s

        from handlers.start import clear_welcome_messages, register_welcome_messages
        clear_welcome_messages(user_id)

        box_text = (
            f"<b>[ PROSES SELESAI ]</b>\n"
            f"<blockquote>"
            f"• Format Output : {fmt.upper()}\n"
            f"• Total Nomor : {len(numbers)}\n"
            f"• Nama File : {file_name}.{fmt}</blockquote>\n\n"
            f"<i>Input manual selesai! Silakan unduh file di atas.</i>"
        )

        final_msg = await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text=box_text,
            parse_mode="HTML",
            reply_markup=keyboard
        )
        register_welcome_messages(user_id, [final_msg.message_id])

    except Exception as e:
        logger.error("Error di proses manual: %s", e, exc_info=True)
        if status_msg_id:
            try:
                await context.bot.edit_message_text(
                    chat_id=update.effective_chat.id,
                    message_id=status_msg_id,
                    text="<blockquote>⚠️ <b>Terjadi kesalahan saat memproses data.</b></blockquote>",
                    parse_mode="HTML"
                )
            except Exception:
                pass

async def handle_show_manual_help_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    asyncio.create_task(adb.increment_usage(user_id))
    
    init_data = {"numbers": []}
    db.set_session(user_id, S_WAIT_TEXT, init_data)
    text = _waiting_text(init_data)
    markup = InlineKeyboardMarkup([[InlineKeyboardButton("BATAL & KEMBALI", callback_data="back_to_start", style="danger")]])

    try:
        await query.message.edit_text(
            text=text,
            parse_mode="HTML",
            reply_markup=markup
        )
        sess = db.get_session(user_id)
        sess["data"]["status_msg_id"] = query.message.message_id
        db.set_session(user_id, S_WAIT_TEXT, sess["data"])
    except Exception:
        try:
            await query.message.delete()
        except Exception:
            pass
        msg = await context.bot.send_message(
            chat_id=query.message.chat_id,
            text=text,
            parse_mode="HTML",
            reply_markup=markup
        )
        sess = db.get_session(user_id)
        sess["data"]["status_msg_id"] = msg.message_id
        db.set_session(user_id, S_WAIT_TEXT, sess["data"])
