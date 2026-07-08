import os
import re
import csv
import logging
import asyncio
from io import BytesIO
from telegram import Update, ReplyKeyboardRemove, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ContextTypes
from database import db
from database.db_async import adb
from middleware.session import get_user_dir

logger = logging.getLogger(__name__)

STATE = "XLSX2TXT_COLLECTING"

MAX_FILES = 100
MAX_SIZE_MB = 50

def _fit(val, max_len=22) -> str:
    s = str(val)
    if len(s) > max_len:
        return s[:max_len-3] + "..."
    return s

_master_locks: dict[int, asyncio.Lock] = {}
_debounce_tasks: dict[int, asyncio.Task] = {}

DEBOUNCE_SECONDS = 1.2

PHONE_REGEX = re.compile(r'\+?(?:\d[\s\-\(\)\.]*){8,16}')


def _get_master_lock(user_id: int) -> asyncio.Lock:
    if user_id not in _master_locks:
        _master_locks[user_id] = asyncio.Lock()
    return _master_locks[user_id]


def cleanup_inactive_users(inactive_ids: list) -> int:
    for uid in inactive_ids:
        _master_locks.pop(uid, None)
        task = _debounce_tasks.pop(uid, None)
        if task:
            task.cancel()
    return len(inactive_ids)


def _get_breadcrumbs(data: dict, step: int) -> str:
    total_file = data.get("total_file", 0)
    total_kontak = data.get("total_kontak", 0)

    parts = []
    if step == 1:
        parts.append(
            f"<b>[UPLOAD BERKAS: {total_file} FILE]</b>" if total_file
            else "<b>[UPLOAD BERKAS]</b>"
        )
    else:
        parts.append(f"Berkas: <code>{total_file}</code>" if total_file else "Berkas: ➖")

    if step == 2:
        parts.append("<b>[SELESAI]</b>")
    else:
        parts.append("Selesai: ➖")

    breadcrumbs = " ➔ ".join(parts)
    return (
        "<b>[ EXCEL/CSV ➔ TXT CONSOLE ]</b>\n"
        "────────────────────────────\n"
        f"<blockquote>{breadcrumbs}</blockquote>\n"
        "────────────────────────────\n\n"
    )


def _waiting_text(data: dict) -> str:
    return (
        _get_breadcrumbs(data, 1) +
        f"<blockquote><b>[ STATUS: WAITING FOR UPLOAD ]</b>\n"
        f"Silakan kirim satu atau beberapa file <code>.xlsx</code> atau <code>.csv</code> sekarang.\n\n"
        f"<b>Batas Sesi:</b>\n"
        f"• Maksimum upload: <code>{MAX_FILES} file</code>\n"
        f"• Maksimum ukuran: <code>{MAX_SIZE_MB} MB</code> per file</blockquote>"
    )


def _extract_numbers_sync(filepath: str, ext: str) -> list:
    numbers = []
    seen = set()
    try:
        def process_cell(cell_value):
            if cell_value is None:
                return
            if isinstance(cell_value, float):
                if cell_value.is_integer():
                    cell_value = int(cell_value)
                else:
                    cell_value = f"{cell_value:.0f}"
            text = str(cell_value).strip()
            for m in PHONE_REGEX.findall(text):
                clean_num = re.sub(r'[^0-9]', '', m)
                if clean_num.startswith("08"):
                    clean_num = "62" + clean_num[1:]
                if 8 <= len(clean_num) <= 15 and clean_num not in seen:
                    seen.add(clean_num)
                    numbers.append(clean_num)

        if ext == ".csv":
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                for row in csv.reader(f):
                    for cell in row:
                        process_cell(cell)
        else:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
            for sheet in wb.sheetnames:
                for row in wb[sheet].iter_rows(values_only=True):
                    for cell in row:
                        process_cell(cell)
            wb.close()
    except Exception as e:
        logger.error("Error ekstrak %s: %s", filepath, e)
    return numbers


def _cancel_debounce(user_id: int):
    old = _debounce_tasks.pop(user_id, None)
    if old and not old.done():
        old.cancel()


def _schedule_debounce(user_id: int, chat_id: int, bot):
    _cancel_debounce(user_id)

    async def _wait_then_notify():
        await asyncio.sleep(DEBOUNCE_SECONDS)
        sess = db.get_session(user_id)
        if not sess or sess.get("state") != STATE:
            return
        data = sess["data"]
        jumlah = data.get("total_file", 0)
        kontak = data.get("total_kontak", 0)

        text = (
            _get_breadcrumbs(data, 1) +
            f"<blockquote><b>[ STATUS: BERKAS DITERIMA ]</b>\n"
            f"Berhasil mengunduh <code>{jumlah}</code> berkas ({kontak:,} nomor unik).\n\n"
            f"Silakan pilih tindakan di bawah:</blockquote>"
        )
        keyboard = InlineKeyboardMarkup([
            [
                InlineKeyboardButton("PROSES SEKARANG", callback_data="done", style="success"),
                InlineKeyboardButton("BATAL & KEMBALI", callback_data="back_to_start", style="danger")
            ]
        ])

        # Hapus welcome messages lama lalu kirim status baru di bawah berkas
        from handlers.start import _welcome_messages
        welcome_ids = _welcome_messages.pop(user_id, [])
        for w_id in welcome_ids:
            try:
                await bot.delete_message(chat_id=chat_id, message_id=w_id)
            except Exception:
                pass

        status_msg_id = data.get("status_msg_id")
        if status_msg_id:
            try:
                await bot.delete_message(chat_id=chat_id, message_id=status_msg_id)
            except Exception:
                pass

        try:
            new_msg = await bot.send_message(
                chat_id=chat_id,
                text=text,
                reply_markup=keyboard,
                parse_mode="HTML"
            )
            data["status_msg_id"] = new_msg.message_id
            db.set_session(user_id, STATE, data)
        except Exception:
            pass

    task = asyncio.create_task(_wait_then_notify())
    _debounce_tasks[user_id] = task


# ─────────────────────────────────────────────────────────────────────────────

async def cmd_xlsxtotxt(update: Update, context: ContextTypes.DEFAULT_TYPE):
    from middleware.auth import require_member
    if not await require_member(update, context):
        return

    user_id = update.effective_user.id
    asyncio.create_task(adb.increment_usage(user_id))

    from handlers.start import transition_to_handler
    import shutil
    _cancel_debounce(user_id)
    xlsx_dir = os.path.join(get_user_dir(user_id), "xlsxtotxt")
    shutil.rmtree(xlsx_dir, ignore_errors=True)
    os.makedirs(xlsx_dir, exist_ok=True)

    master_txt = os.path.join(xlsx_dir, "extracted_numbers.txt")
    open(master_txt, 'w').close()

    init_data = {"total_kontak": 0, "total_file": 0}
    db.set_session(user_id, STATE, init_data)

    msg = await transition_to_handler(
        context.bot,
        user_id,
        update.effective_chat.id,
        _waiting_text(init_data),
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("BATAL & KEMBALI", callback_data="back_to_start", style="danger")]]),
        update=update
    )

    if msg:
        sess = db.get_session(user_id)
        sess["data"]["status_msg_id"] = msg.message_id
        db.set_session(user_id, STATE, sess["data"])


async def handle_xlsxtotxt_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    chat_id = update.effective_chat.id
    sess = db.get_session(user_id)
    if not sess or sess.get("state") != STATE:
        return

    doc = update.message.document
    ext = os.path.splitext(doc.file_name)[1].lower() if doc and doc.file_name else ""
    if not doc or not doc.file_name or ext not in (".xlsx", ".csv"):
        # Format salah — edit status message in-place, JANGAN hapus file user
        try:
            status_msg_id = sess["data"].get("status_msg_id")
            sent_name = doc.file_name if doc and doc.file_name else "file tersebut"
            if status_msg_id:
                await context.bot.edit_message_text(
                    chat_id=chat_id,
                    message_id=status_msg_id,
                    text=(
                        _get_breadcrumbs(sess["data"], 1) +
                        f"<blockquote>⚠️ <b>[ FORMAT SALAH ]</b>\n"
                        f"<code>{sent_name}</code> bukan berkas <code>.xlsx</code> atau <code>.csv</code>.</blockquote>"
                    ),
                    parse_mode="HTML",
                    reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("BATAL & KEMBALI", callback_data="back_to_start", style="danger")]])
                )
                await asyncio.sleep(10)
                await context.bot.edit_message_text(
                    chat_id=chat_id,
                    message_id=status_msg_id,
                    text=_waiting_text(sess["data"]),
                    parse_mode="HTML",
                    reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("BATAL & KEMBALI", callback_data="back_to_start", style="danger")]])
                )
        except Exception:
            pass
        return

    xlsx_dir = os.path.join(get_user_dir(user_id), "xlsxtotxt")
    os.makedirs(xlsx_dir, exist_ok=True)
    file_path = os.path.join(xlsx_dir, f"{doc.file_id}{ext}")

    try:
        tg_file = await context.bot.get_file(doc.file_id)
        await tg_file.download_to_drive(file_path)
    except Exception as e:
        logger.error("Download error user %s: %s", user_id, e)
        try:
            status_msg_id = sess["data"].get("status_msg_id")
            if status_msg_id:
                await context.bot.edit_message_text(
                    chat_id=chat_id,
                    message_id=status_msg_id,
                    text=(
                        _get_breadcrumbs(sess["data"], 1) +
                        f"<blockquote>⚠️ <b>[ GAGAL UNDUH ]</b>\n"
                        f"Gagal mengunduh <code>{doc.file_name}</code>. Coba kirim ulang.</blockquote>"
                    ),
                    parse_mode="HTML",
                    reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("BATAL & KEMBALI", callback_data="back_to_start", style="danger")]])
                )
        except Exception:
            pass
        return

    loop = asyncio.get_running_loop()
    found_numbers = await loop.run_in_executor(None, _extract_numbers_sync, file_path, ext)

    try:
        os.remove(file_path)
    except Exception:
        pass

    master_txt = os.path.join(xlsx_dir, "extracted_numbers.txt")
    new_total = 0

    async with _get_master_lock(user_id):
        try:
            with open(master_txt, 'r', encoding='utf-8') as f:
                existing = f.read().splitlines()
        except FileNotFoundError:
            existing = []

        seen = set(existing)
        combined = list(existing)
        for num in found_numbers:
            if num not in seen:
                seen.add(num)
                combined.append(num)

        with open(master_txt, 'w', encoding='utf-8') as f:
            f.write("\n".join(combined))

        new_total = len(combined)

        fresh = db.get_session(user_id)
        if fresh and fresh.get("state") == STATE:
            data = fresh["data"]
            data["total_file"] += 1
            data["total_kontak"] = new_total
            db.set_session(user_id, STATE, data)

    # Pindahkan status ke bawah berkas (debounce)
    _schedule_debounce(user_id, chat_id, context.bot)


async def handle_xlsxtotxt_done(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    _cancel_debounce(user_id)

    sess = db.get_session(user_id)
    if not sess or sess.get("state") != STATE:
        return

    # Hapus input teks 'done' jika user mengetik manual
    if update.message and update.message.text in ("done", "selesai", "/done"):
        try:
            await update.message.delete()
        except Exception:
            pass

    data = sess["data"]
    total = data.get("total_kontak", 0)
    total_file = data.get("total_file", 0)
    status_msg_id = data.get("status_msg_id")

    # ── Edit status message in-place: Processing ────────────────────────────
    process_text = "<blockquote><b>[ SYSTEM: PROCESSING DATA ]</b>\nSedang mengekstrak nomor dari berkas Excel/CSV...</blockquote>"
    if update.callback_query:
        try:
            await update.callback_query.message.edit_text(
                text=process_text,
                parse_mode="HTML"
            )
        except Exception:
            pass
    elif status_msg_id:
        try:
            await context.bot.edit_message_text(
                chat_id=update.effective_chat.id,
                message_id=status_msg_id,
                text=process_text,
                parse_mode="HTML"
            )
        except Exception:
            pass

    keyboard = InlineKeyboardMarkup([
        [
            InlineKeyboardButton("PROSES FILE LAIN", callback_data="show_xlsxtotxt_help", style="success"),
            InlineKeyboardButton("KEMBALI KE MENU", callback_data="back_to_start", style="danger")
        ]
    ])

    if total == 0:
        from handlers.start import clear_welcome_messages, register_welcome_messages
        clear_welcome_messages(user_id)

        box_empty = (
            _get_breadcrumbs(data, 2) +
            "<blockquote>⚠️ <b>[ TIDAK ADA NOMOR ]</b>\n"
            "Tidak ada nomor yang ditemukan dari berkas yang dikirim.</blockquote>"
        )
        if update.callback_query:
            try:
                await update.callback_query.message.edit_text(
                    text=box_empty,
                    parse_mode="HTML",
                    reply_markup=keyboard
                )
                final_id = update.callback_query.message.message_id
            except Exception:
                final_msg = await context.bot.send_message(
                    chat_id=update.effective_chat.id,
                    text=box_empty,
                    parse_mode="HTML",
                    reply_markup=keyboard
                )
                final_id = final_msg.message_id
        else:
            if status_msg_id:
                try:
                    await context.bot.edit_message_text(
                        chat_id=update.effective_chat.id,
                        message_id=status_msg_id,
                        text=box_empty,
                        parse_mode="HTML",
                        reply_markup=keyboard
                    )
                    final_id = status_msg_id
                except Exception:
                    final_msg = await context.bot.send_message(
                        chat_id=update.effective_chat.id,
                        text=box_empty,
                        parse_mode="HTML",
                        reply_markup=keyboard
                    )
                    final_id = final_msg.message_id
            else:
                final_msg = await context.bot.send_message(
                    chat_id=update.effective_chat.id,
                    text=box_empty,
                    parse_mode="HTML",
                    reply_markup=keyboard
                )
                final_id = final_msg.message_id

        register_welcome_messages(user_id, [final_id])
        db.clear_session(user_id)
        import shutil
        shutil.rmtree(os.path.join(get_user_dir(user_id), "xlsxtotxt"), ignore_errors=True)
        return

    xlsx_dir = os.path.join(get_user_dir(user_id), "xlsxtotxt")
    master_txt = os.path.join(xlsx_dir, "extracted_numbers.txt")

    try:
        with open(master_txt, 'rb') as f:
            buffer = BytesIO(f.read())
            buffer.name = "Hasil_Ekstrak.txt"

        # Kirim berkas hasil
        await context.bot.send_document(
            chat_id=update.effective_chat.id,
            document=buffer,
            filename="Hasil_Ekstrak.txt",
            caption=(
                f"<b>File hasil ekstraksi</b>\n"
                f"Berkas: <code>{total_file}</code> | Nomor unik: <code>{total:,}</code>"
            ),
            parse_mode="HTML",
        )

        from handlers.start import clear_welcome_messages, register_welcome_messages
        clear_welcome_messages(user_id)

        box_text = (
            f"<b>[ PROSES SELESAI ]</b>\n"
            f"<blockquote>"
            f"• Total Berkas : {total_file} EXCEL/CSV\n"
            f"• Total Kontak : {total:,}</blockquote>\n\n"
            f"<i>Silakan unduh file hasil ekstraksi di atas.</i>"
        )

        # Hapus status message lama — kirim laporan sukses di bawah berkas
        if status_msg_id:
            try:
                await context.bot.delete_message(
                    chat_id=update.effective_chat.id,
                    message_id=status_msg_id
                )
            except Exception:
                pass

        final_msg = await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text=box_text,
            parse_mode="HTML",
            reply_markup=keyboard
        )
        register_welcome_messages(user_id, [final_msg.message_id])

    except Exception as e:
        logger.error("Error kirim hasil xlsx: %s", e)
        try:
            if status_msg_id:
                await context.bot.edit_message_text(
                    chat_id=update.effective_chat.id,
                    message_id=status_msg_id,
                    text="<blockquote>⚠️ <b>Gagal mengirim hasil. Coba ulangi.</b></blockquote>",
                    parse_mode="HTML"
                )
        except Exception:
            pass
    finally:
        db.clear_session(user_id)
        import shutil
        shutil.rmtree(xlsx_dir, ignore_errors=True)


async def handle_show_xlsxtotxt_help_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Callback untuk tombol PROSES FILE LAIN (Excel/CSV to TXT)"""
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    asyncio.create_task(adb.increment_usage(user_id))

    import shutil
    _cancel_debounce(user_id)
    xlsx_dir = os.path.join(get_user_dir(user_id), "xlsxtotxt")
    shutil.rmtree(xlsx_dir, ignore_errors=True)
    os.makedirs(xlsx_dir, exist_ok=True)

    master_txt = os.path.join(xlsx_dir, "extracted_numbers.txt")
    open(master_txt, 'w').close()

    init_data = {"total_kontak": 0, "total_file": 0}
    db.set_session(user_id, STATE, init_data)
    text = _waiting_text(init_data)
    markup = InlineKeyboardMarkup([[InlineKeyboardButton("BATAL & KEMBALI", callback_data="back_to_start", style="danger")]])

    try:
        await query.message.edit_text(
            text=text,
            parse_mode="HTML",
            reply_markup=markup
        )
        sess = db.get_session(user_id)
        sess["data"]["status_msg_id"] = query.message.message_id
        db.set_session(user_id, STATE, sess["data"])
    except Exception:
        try:
            await query.message.delete()
        except Exception:
            pass
        msg = await context.bot.send_message(
            chat_id=query.message.chat_id,
            text=text,
            parse_mode="HTML",
            reply_markup=markup
        )
        sess = db.get_session(user_id)
        sess["data"]["status_msg_id"] = msg.message_id
        db.set_session(user_id, STATE, sess["data"])