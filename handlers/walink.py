"""
handlers/walink.py — Pembuat tautan WhatsApp instan dalam format Excel (.xlsx) premium.
Mendukung berkas input .xlsx, .csv, .txt, dan .vcf.
"""
import os
import io
import re
import csv
import stat
import shutil
import asyncio
import logging
import openpyxl
import urllib.parse
from openpyxl.styles import Font, Alignment, PatternFill
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ContextTypes
from database import db
from database.db_async import adb
from middleware.auth import require_member
from middleware.session import get_user_dir
from core.vcf_parser import parse_vcf_file

logger = logging.getLogger(__name__)

S0 = "WALINK_WAIT_FILE"
S1 = "WALINK_WAIT_MSG"

_processing: set[int] = set()
_button_timers: dict[int, asyncio.Task] = {}
_user_locks: dict = {}

def cleanup_inactive_users(inactive_ids: list) -> int:
    cleaned = 0
    for uid in inactive_ids:
        _processing.discard(uid)
        _user_locks.pop(uid, None)
        task = _button_timers.pop(uid, None)
        if task:
            task.cancel()
        cleaned += 1
    return cleaned

def _get_lock(user_id: int) -> asyncio.Lock:
    return _user_locks.setdefault(user_id, asyncio.Lock())

PHONE_REGEX = re.compile(r'\+?(?:\d[\s\-\(\)\.]*){8,16}')

def _clean_number(num: str) -> str:
    """Bersihkan nomor ke format standard 628xxx."""
    digits = re.sub(r'\D', '', num)
    if digits.startswith("08"):
        digits = "62" + digits[1:]
    elif digits.startswith("8"):
        digits = "62" + digits
    if 9 <= len(digits) <= 15:
        return digits
    return ""

def _extract_contacts_sync(filepath: str, ext: str) -> list:
    contacts = []
    seen = set()
    
    def process_number(num: str, name: str = None):
        clean = _clean_number(num)
        if clean and clean not in seen:
            seen.add(clean)
            contacts.append({
                "name": name or f"Kontak {len(contacts) + 1}",
                "tel": clean
            })

    try:
        if ext == ".vcf":
            parsed = parse_vcf_file(filepath)
            for c in parsed:
                process_number(c["tel"], c["name"])
        elif ext == ".csv":
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                for row in csv.reader(f):
                    for cell in row:
                        if cell:
                            text = str(cell).strip()
                            for m in PHONE_REGEX.findall(text):
                                process_number(m)
        elif ext == ".xlsx":
            wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
            for sheet in wb.sheetnames:
                for row in wb[sheet].iter_rows(values_only=True):
                    for cell in row:
                        if cell:
                            text = str(cell).strip()
                            for m in PHONE_REGEX.findall(text):
                                process_number(m)
            wb.close()
        else:
            # TXT file
            with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
                for line in f:
                    stripped = line.strip()
                    if stripped:
                        for m in PHONE_REGEX.findall(stripped):
                            process_number(m)
    except Exception as e:
        logger.error("Error extracting contacts in walink: %s", e)
    return contacts

def _get_breadcrumbs(data: dict, step: int) -> str:
    count = data.get("count", 0)
    custom_msg = data.get("custom_msg", "")
    parts = []
    
    # Step 1: Berkas
    if step == 1:
        parts.append(f"<b>» BERKAS: {count} FILE «</b>" if count else "<b>» BERKAS «</b>")
    else:
        parts.append(f"Berkas: {count} file" if count else "Berkas ○")
        
    # Step 2: Pesan Kustom
    if step == 2:
        parts.append(f"<b>» PESAN: {custom_msg[:12].upper()}... «</b>" if custom_msg else "<b>» PESAN «</b>")
    elif step > 2:
        parts.append(f"Pesan: {custom_msg[:8]}..." if custom_msg else "Pesan -")
    else:
        parts.append("Pesan ○")
        
    # Step 3: Excel
    if step == 3:
        parts.append("<b>» EXCEL «</b>")
    else:
        parts.append("Excel ○")
        
    breadcrumbs = " ➔ ".join(parts)
    return (
        "<b>[ WALINK EXCEL CV ]</b>\n"
        f"{breadcrumbs}\n"
        "\n"
    )

async def cmd_walink(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await require_member(update, context):
        return
    user_id = update.effective_user.id
    asyncio.create_task(adb.increment_usage(user_id))

    db.set_session(user_id, S0, {"count": 0, "total_size": 0})
    from handlers.start import transition_to_handler
    
    text = _get_breadcrumbs({"count": 0}, 1) + "<b>[ ➔ ] Menunggu berkas...</b>\nKirim file <b>.xlsx</b>, <b>.csv</b>, <b>.txt</b>, atau <b>.vcf</b> sekarang."
    
    msg = await transition_to_handler(
        context.bot,
        user_id,
        update.effective_chat.id,
        text,
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("BATAL & KEMBALI", callback_data="back_to_start", style="danger")]]),
        update=update
    )
    if msg:
        sess = db.get_session(user_id)
        sess["data"]["status_msg_id"] = msg.message_id
        db.set_session(user_id, S0, sess["data"])

async def handle_walink_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    sess = db.get_session(user_id)
    if not sess or sess.get("state") != S0:
        return

    doc = update.message.document
    try:
        await update.message.delete()
    except Exception:
        pass

    data = sess["data"]
    status_msg_id = data.get("status_msg_id")
    keyboard = InlineKeyboardMarkup([[InlineKeyboardButton("BATAL & KEMBALI", callback_data="back_to_start", style="danger")]])

    if not doc or not doc.file_name:
        try:
            await context.bot.edit_message_text(
                chat_id=update.effective_chat.id,
                message_id=status_msg_id,
                text=_get_breadcrumbs(data, 1) + "Kirim file dokumen valid (.xlsx, .csv, .txt, .vcf).",
                parse_mode="HTML",
                reply_markup=keyboard
            )
        except Exception:
            pass
        return

    ext = os.path.splitext(doc.file_name)[1].lower()
    if ext not in (".xlsx", ".csv", ".txt", ".vcf"):
        try:
            await context.bot.edit_message_text(
                chat_id=update.effective_chat.id,
                message_id=status_msg_id,
                text=_get_breadcrumbs(data, 1) + "Format tidak didukung. Kirim file .xlsx, .csv, .txt, atau .vcf.",
                parse_mode="HTML",
                reply_markup=keyboard
            )
        except Exception:
            pass
        return

    async with _get_lock(user_id):
        if user_id in _processing:
            return

    user_dir = get_user_dir(user_id)
    work_dir = os.path.join(user_dir, f"walink_{doc.file_id}")

    try:
        os.makedirs(work_dir, exist_ok=True)
        input_path = os.path.join(work_dir, f"input{ext}")
        file_obj = await context.bot.get_file(doc.file_id)
        await file_obj.download_to_drive(input_path)
        
        data.update({
            "input_path": input_path,
            "ext": ext,
            "file_name": doc.file_name,
            "work_dir": work_dir,
            "count": 1
        })
        db.set_session(user_id, S1, data)
        
        try:
            await context.bot.edit_message_text(
                chat_id=update.effective_chat.id,
                message_id=status_msg_id,
                text=_get_breadcrumbs(data, 2) + "Tulis pesan WhatsApp kustom untuk tautan follow-up?\nKetik isi pesannya sekarang, atau ketik <b>-</b> (tanda minus) untuk mengosongkan tanpa pesan.",
                parse_mode="HTML",
                reply_markup=keyboard
            )
        except Exception:
            pass
    except Exception as e:
        logger.error("Failed to download file in walink: %s", e)
        asyncio.create_task(asyncio.to_thread(shutil.rmtree, work_dir, ignore_errors=True))
        try:
            await context.bot.edit_message_text(
                chat_id=update.effective_chat.id,
                message_id=status_msg_id,
                text=_get_breadcrumbs(data, 1) + "Gagal mengunduh berkas. Coba kirim ulang.",
                parse_mode="HTML",
                reply_markup=keyboard
            )
        except Exception:
            pass

async def handle_walink_msg(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    sess = db.get_session(user_id)
    if not sess or sess.get("state") != S1:
        return

    msg_text = update.message.text.strip()
    try:
        await update.message.delete()
    except Exception:
        pass

    custom_msg = "" if msg_text == "-" else msg_text

    data = sess["data"]
    input_path = data["input_path"]
    ext = data["ext"]
    orig_name = data["file_name"]
    work_dir = data["work_dir"]
    status_msg_id = data.get("status_msg_id")

    async with _get_lock(user_id):
        if user_id in _processing:
            return
        _processing.add(user_id)

    data["custom_msg"] = custom_msg
    db.set_session(user_id, S1, data)

    try:
        await context.bot.edit_message_text(
            chat_id=update.effective_chat.id,
            message_id=status_msg_id,
            text=_get_breadcrumbs(data, 3) + "<b>Memproses berkas dan membuat Excel...</b>",
            parse_mode="HTML"
        )
    except Exception:
        pass

    keyboard = InlineKeyboardMarkup([[InlineKeyboardButton("BATAL & KEMBALI", callback_data="back_to_start", style="danger")]])

    try:
        loop = asyncio.get_running_loop()

        def do_process_xlsx():
            contacts = _extract_contacts_sync(input_path, ext)
            if not contacts:
                return 0, None

            # Membuat Workbook Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "WhatsApp Links"

            # Style Header
            header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center")

            headers = ["No", "Nama Kontak", "Nomor HP", "Link WhatsApp"]
            ws.append(headers)

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            # Isi Baris
            row_align_center = Alignment(horizontal="center", vertical="center")
            row_align_left = Alignment(horizontal="left", vertical="center")
            
            for idx, c in enumerate(contacts, 1):
                if custom_msg:
                    quoted_text = urllib.parse.quote(custom_msg)
                    wa_link = f"https://wa.me/{c['tel']}?text={quoted_text}"
                else:
                    wa_link = f"https://wa.me/{c['tel']}"
                ws.append([idx, c["name"], c["tel"], wa_link])
                
                # Formatting link biar bisa langsung diklik sebagai hyperlink di excel
                cell_link = ws.cell(row=idx+1, column=4)
                cell_link.hyperlink = wa_link
                cell_link.font = Font(name="Arial", size=10, color="0000FF", underline="single")
                
                # Alignment
                ws.cell(row=idx+1, column=1).alignment = row_align_center
                ws.cell(row=idx+1, column=2).alignment = row_align_left
                ws.cell(row=idx+1, column=3).alignment = row_align_center
                ws.cell(row=idx+1, column=4).alignment = row_align_left

            # Atur Lebar Kolom secara otomatis
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return len(contacts), buf

        total_links, excel_buf = await loop.run_in_executor(None, do_process_xlsx)

        if total_links == 0:
            try:
                await context.bot.edit_message_text(
                    chat_id=update.effective_chat.id,
                    message_id=status_msg_id,
                    text=_get_breadcrumbs(data, 1) + "Tidak ada nomor HP valid yang ditemukan dalam file.",
                    parse_mode="HTML",
                    reply_markup=keyboard
                )
            except Exception:
                pass
            db.clear_session(user_id)
            return

        # Kirim file Excel kembali ke user
        base_name = os.path.splitext(orig_name)[0]
        out_name = f"WA_LINKS_{base_name}.xlsx"
        excel_buf.name = out_name

        await context.bot.send_document(
            chat_id=update.effective_chat.id,
            document=excel_buf,
            filename=out_name
        )

        try:
            await context.bot.delete_message(chat_id=update.effective_chat.id, message_id=status_msg_id)
        except Exception:
            pass

        old_timer = _button_timers.pop(user_id, None)
        if old_timer and not old_timer.done():
            old_timer.cancel()

        async def _send_buttons_debounced(uid, chat_id, bot):
            await asyncio.sleep(1.5)
            keyboard_done = InlineKeyboardMarkup([
                [
                    InlineKeyboardButton("PROSES FILE LAIN", callback_data="show_walink_help", style="success"),
                    InlineKeyboardButton("KEMBALI KE MENU", callback_data="back_to_start", style="danger")
                ]
            ])

            from handlers.start import clear_welcome_messages, register_welcome_messages
            clear_welcome_messages(uid)

            box_text = (
                f"<b>[ PROSES SELESAI ]</b>\n"
                f"<blockquote>"
                f"• File Input : {orig_name}\n"
                f"• Berkas Output : 1 EXCEL (.xlsx)\n"
                f"• Total Link WA : {total_links:,}</blockquote>\n\n"
                f"<i>Pembuatan WA Link Excel selesai! Silakan unduh file di atas.</i>"
            )
            
            # Fallback jika edit gagal
            final_msg = await bot.send_message(
                chat_id=chat_id,
                text=box_text,
                parse_mode="HTML",
                reply_markup=keyboard_done
            )
            register_welcome_messages(uid, [final_msg.message_id])
            db.clear_session(uid)

        task = asyncio.create_task(_send_buttons_debounced(user_id, update.effective_chat.id, context.bot))
        _button_timers[user_id] = task

    except Exception as e:
        logger.error("Error di walink: %s", e, exc_info=True)
        try:
            await context.bot.edit_message_text(
                chat_id=update.effective_chat.id,
                message_id=status_msg_id,
                text="Terjadi kesalahan saat memproses file. Coba lagi.",
                parse_mode="HTML",
                reply_markup=keyboard
            )
        except Exception:
            pass
    finally:
        async with _get_lock(user_id):
            _processing.discard(user_id)
        asyncio.create_task(asyncio.to_thread(shutil.rmtree, work_dir, ignore_errors=True))

async def handle_show_walink_help_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    db.set_session(user_id, S0, {"count": 0, "total_size": 0})

    text = _get_breadcrumbs({"count": 0}, 1) + "<b>[ ➔ ] Menunggu berkas...</b>\nKirim file <b>.xlsx</b>, <b>.csv</b>, <b>.txt</b>, atau <b>.vcf</b> sekarang."

    try:
        await query.message.edit_text(
            text=text,
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("BATAL & KEMBALI", callback_data="back_to_start", style="danger")]])
        )
        sess = db.get_session(user_id)
        sess["data"]["status_msg_id"] = query.message.message_id
        db.set_session(user_id, S0, sess["data"])
    except Exception:
        try:
            await query.message.delete()
        except Exception:
            pass
        msg = await context.bot.send_message(
            chat_id=query.message.chat_id,
            text=text,
            parse_mode="HTML",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("BATAL & KEMBALI", callback_data="back_to_start", style="danger")]])
        )
        sess = db.get_session(user_id)
        sess["data"]["status_msg_id"] = msg.message_id
        db.set_session(user_id, S0, sess["data"])
