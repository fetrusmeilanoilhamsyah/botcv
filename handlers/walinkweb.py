"""
handlers/walinkweb.py — Pembuat dashboard tautan WhatsApp interaktif premium dalam format HTML.
Mendukung berkas input .TXT, .VCF, .XLSX, dan .CSV.
"""
import os
import re
import csv
import json
import shutil
import io
import urllib.parse
import asyncio
import logging
from telegram import Update, ReplyKeyboardRemove, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ContextTypes
from database import db
from database.db_async import adb
from middleware.auth import require_member
from middleware.session import get_user_dir
from core.vcf_parser import parse_vcf_file

logger = logging.getLogger(__name__)

S0 = "WALINKWEB_WAIT_FILE"
S1 = "WALINKWEB_WAIT_MSG"
_processing: set[int] = set()
_button_timers: dict[int, asyncio.Task] = {}
_user_locks: dict = {}


def cleanup_inactive_users(inactive_ids: list) -> int:
    cleaned = 0
    for uid in inactive_ids:
        _processing.discard(uid)
        _user_locks.pop(uid, None)
        task = _button_timers.pop(uid, None)
        if task:
            task.cancel()
        cleaned += 1
    return cleaned

def _get_lock(user_id: int) -> asyncio.Lock:
    return _user_locks.setdefault(user_id, asyncio.Lock())

PHONE_REGEX = re.compile(r'\+?(?:\d[\s\-\(\)\.]*){8,16}')


def _clean_number(num: str) -> str:
    """Bersihkan nomor ke format standard 628xxx."""
    digits = re.sub(r'\D', '', num)
    if digits.startswith("08"):
        digits = "62" + digits[1:]
    elif digits.startswith("8"):
        digits = "62" + digits
    if 9 <= len(digits) <= 15:
        return digits
    return ""


def _extract_contacts_sync(filepath: str, ext: str) -> list:
    contacts = []
    seen = set()
    
    def process_number(num: str, name: str = None):
        clean = _clean_number(num)
        if clean and clean not in seen:
            seen.add(clean)
            contacts.append({
                "name": name or f"Kontak {len(contacts) + 1}",
                "tel": clean
            })

    try:
        if ext == ".vcf":
            parsed = parse_vcf_file(filepath)
            for c in parsed:
                process_number(c["tel"], c["name"])
        elif ext == ".csv":
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                for row in csv.reader(f):
                    for cell in row:
                        if cell:
                            text = str(cell).strip()
                            for m in PHONE_REGEX.findall(text):
                                process_number(m)
        elif ext == ".xlsx":
            import openpyxl
            wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
            for sheet in wb.sheetnames:
                for row in wb[sheet].iter_rows(values_only=True):
                    for cell in row:
                        if cell:
                            text = str(cell).strip()
                            for m in PHONE_REGEX.findall(text):
                                process_number(m)
            wb.close()
        else:
            # TXT file
            with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
                for line in f:
                    stripped = line.strip()
                    if stripped:
                        for m in PHONE_REGEX.findall(stripped):
                            process_number(m)
    except Exception as e:
        logger.error("Error extracting contacts in walinkweb: %s", e)
    return contacts


HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="id">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>CS WA Links Dashboard — {filename}</title>
    <link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;600;700&display=swap" rel="stylesheet">
    <style>
        :root {{
            --bg-color: #0d1117;
            --card-bg: rgba(22, 27, 34, 0.75);
            --border-color: rgba(240, 246, 252, 0.1);
            --primary-gradient: linear-gradient(135deg, #10b981, #059669);
            --accent-gradient: linear-gradient(135deg, #3b82f6, #1d4ed8);
            --text-color: #c9d1d9;
            --text-title: #f0f6fc;
            --text-muted: #8b949e;
        }}

        * {{
            box-sizing: border-box;
            margin: 0;
            padding: 0;
            font-family: 'Outfit', sans-serif;
            transition: all 0.2s cubic-bezier(0.4, 0, 0.2, 1);
        }}

        body {{
            background-color: var(--bg-color);
            background-image: radial-gradient(circle at 10% 20%, rgba(16, 185, 129, 0.05) 0%, transparent 40%),
                              radial-gradient(circle at 90% 80%, rgba(59, 130, 246, 0.05) 0%, transparent 40%);
            color: var(--text-color);
            min-height: 100vh;
            display: flex;
            flex-direction: column;
            align-items: center;
            padding: 2rem 1rem;
        }}

        .container {{
            width: 100%;
            max-width: 900px;
            display: flex;
            flex-direction: column;
            gap: 1.5rem;
        }}

        header {{
            text-align: center;
            padding: 1.75rem;
            background: var(--card-bg);
            border: 1px solid var(--border-color);
            backdrop-filter: blur(12px);
            border-radius: 24px;
            box-shadow: 0 8px 32px 0 rgba(0, 0, 0, 0.3);
        }}

        h1 {{
            font-size: 2.2rem;
            font-weight: 700;
            background: linear-gradient(135deg, #10b981, #3b82f6);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            margin-bottom: 0.5rem;
        }}

        .subtitle {{
            font-size: 0.95rem;
            color: var(--text-muted);
        }}

        .stats-bar {{
            display: grid;
            grid-template-columns: repeat(3, 1fr);
            gap: 1rem;
            text-align: center;
        }}

        .stat-card {{
            background: var(--card-bg);
            border: 1px solid var(--border-color);
            backdrop-filter: blur(12px);
            padding: 1.25rem 1rem;
            border-radius: 20px;
            box-shadow: 0 4px 20px 0 rgba(0, 0, 0, 0.15);
        }}

        .stat-val {{
            font-size: 2rem;
            font-weight: 700;
            color: var(--text-title);
        }}

        .stat-label {{
            font-size: 0.8rem;
            color: var(--text-muted);
            text-transform: uppercase;
            letter-spacing: 0.05em;
            margin-top: 0.25rem;
        }}

        .controls {{
            display: flex;
            gap: 1rem;
            width: 100%;
        }}

        .search-wrapper {{
            position: relative;
            flex-grow: 1;
        }}

        .search-input {{
            width: 100%;
            padding: 0.85rem 1rem 0.85rem 2.75rem;
            background: var(--card-bg);
            border: 1px solid var(--border-color);
            border-radius: 16px;
            color: var(--text-title);
            font-size: 1rem;
            outline: none;
            backdrop-filter: blur(12px);
        }}

        .search-input:focus {{
            border-color: #10b981;
            box-shadow: 0 0 12px rgba(16, 185, 129, 0.25);
        }}

        .search-icon {{
            position: absolute;
            left: 1rem;
            top: 50%;
            transform: translateY(-50%);
            color: var(--text-muted);
            pointer-events: none;
            font-size: 1.1rem;
        }}

        .filter-btn {{
            padding: 0.85rem 1.5rem;
            background: var(--card-bg);
            border: 1px solid var(--border-color);
            border-radius: 16px;
            color: var(--text-title);
            font-size: 0.95rem;
            font-weight: 600;
            cursor: pointer;
            backdrop-filter: blur(12px);
            display: flex;
            align-items: center;
            gap: 0.5rem;
            white-space: nowrap;
        }}

        .filter-btn:hover {{
            border-color: #3b82f6;
            background: rgba(59, 130, 246, 0.1);
        }}

        .filter-btn.active {{
            background: var(--accent-gradient);
            border-color: transparent;
            box-shadow: 0 4px 12px rgba(59, 130, 246, 0.25);
        }}

        .contact-list {{
            display: flex;
            flex-direction: column;
            gap: 0.75rem;
            max-height: 550px;
            overflow-y: auto;
            padding-right: 0.25rem;
        }}

        .contact-list::-webkit-scrollbar {{
            width: 6px;
        }}
        .contact-list::-webkit-scrollbar-track {{
            background: transparent;
        }}
        .contact-list::-webkit-scrollbar-thumb {{
            background: rgba(255, 255, 255, 0.1);
            border-radius: 3px;
        }}
        .contact-list::-webkit-scrollbar-thumb:hover {{
            background: rgba(255, 255, 255, 0.2);
        }}

        .contact-card {{
            background: var(--card-bg);
            border: 1px solid var(--border-color);
            border-radius: 18px;
            padding: 1.1rem 1.25rem;
            display: flex;
            align-items: center;
            justify-content: space-between;
            box-shadow: 0 4px 15px 0 rgba(0, 0, 0, 0.1);
            backdrop-filter: blur(8px);
        }}

        .contact-card.clicked {{
            opacity: 0.6;
            border-color: rgba(16, 185, 129, 0.3);
            background: rgba(16, 185, 129, 0.02);
        }}

        .contact-info {{
            display: flex;
            flex-direction: column;
            gap: 0.25rem;
        }}

        .contact-name {{
            font-size: 1.1rem;
            font-weight: 600;
            color: var(--text-title);
        }}

        .contact-phone {{
            font-size: 0.9rem;
            color: var(--text-muted);
            letter-spacing: 0.02em;
        }}

        .action-area {{
            display: flex;
            align-items: center;
            gap: 1rem;
        }}

        .status-badge {{
            font-size: 0.75rem;
            font-weight: 700;
            padding: 0.25rem 0.65rem;
            border-radius: 20px;
            background: rgba(255, 255, 255, 0.05);
            color: var(--text-muted);
            border: 1px solid var(--border-color);
            letter-spacing: 0.05em;
        }}

        .contact-card.clicked .status-badge {{
            background: rgba(16, 185, 129, 0.12);
            color: #10b981;
            border-color: rgba(16, 185, 129, 0.2);
        }}

        .wa-btn {{
            background: var(--primary-gradient);
            color: #ffffff;
            border: none;
            padding: 0.65rem 1.2rem;
            border-radius: 12px;
            font-size: 0.9rem;
            font-weight: 600;
            text-decoration: none;
            cursor: pointer;
            display: inline-flex;
            align-items: center;
            gap: 0.4rem;
            box-shadow: 0 4px 12px rgba(16, 185, 129, 0.2);
        }}

        .wa-btn:hover {{
            transform: translateY(-2px);
            box-shadow: 0 6px 16px rgba(16, 185, 129, 0.35);
        }}

        .wa-btn:active {{
            transform: translateY(0);
        }}

        .contact-card.clicked .wa-btn {{
            background: var(--accent-gradient);
            box-shadow: 0 4px 12px rgba(59, 130, 246, 0.2);
        }}

        .empty-state {{
            text-align: center;
            padding: 4rem 2rem;
            color: var(--text-muted);
            background: var(--card-bg);
            border: 1px solid var(--border-color);
            border-radius: 24px;
        }}

        footer {{
            margin-top: auto;
            padding-top: 3rem;
            font-size: 0.85rem;
            color: var(--text-muted);
            text-align: center;
        }}

        @media (max-width: 600px) {{
            .stats-bar {{
                grid-template-columns: 1fr;
                gap: 0.75rem;
            }}
            .controls {{
                flex-direction: column;
                gap: 0.75rem;
            }}
            .filter-btn {{
                justify-content: center;
            }}
            .contact-card {{
                flex-direction: column;
                align-items: flex-start;
                gap: 1rem;
            }}
            .action-area {{
                width: 100%;
                justify-content: space-between;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <header>
            <h1>CS WA Links Dashboard</h1>
            <p class="subtitle" id="date-label">Mempermudah follow-up tim Customer Service & Admin</p>
        </header>

        <div class="stats-bar">
            <div class="stat-card">
                <div class="stat-val" id="total-cnt">0</div>
                <div class="stat-label">Total Nomor</div>
            </div>
            <div class="stat-card">
                <div class="stat-val" id="clicked-cnt" style="color: #10b981;">0</div>
                <div class="stat-label">Sudah Hubungi</div>
            </div>
            <div class="stat-card">
                <div class="stat-val" id="left-cnt" style="color: #3b82f6;">0</div>
                <div class="stat-label">Belum Hubungi</div>
            </div>
        </div>

        <div class="controls">
            <div class="search-wrapper">
                <span class="search-icon">🔍</span>
                <input type="text" class="search-input" id="search-box" placeholder="Cari nama kontak atau nomor HP...">
            </div>
            <button class="filter-btn" id="filter-btn">
                <span>👁️</span> Tampilkan Belum Hubungi
            </button>
        </div>

        <div class="contact-list" id="contact-list">
            <!-- Kontak dimasukkan via JS -->
        </div>
    </div>

    <footer>
        <p>DiBot CV FEE — Generated Premium Web Links Dashboard</p>
    </footer>

    <script>
        const CONTACTS = {contacts_json};
        const PAGE_ID = "{page_id}";

        let showOnlyUnclicked = false;

        const getClickedMap = () => {{
            try {{
                return JSON.parse(localStorage.getItem('wa_clicked_' + PAGE_ID) || '{{}}');
            }} catch {{
                return {{}};
            }}
        }};

        const saveClickedMap = (map) => {{
            localStorage.setItem('wa_clicked_' + PAGE_ID, JSON.stringify(map));
        }};

        const render = () => {{
            const list = document.getElementById('contact-list');
            const search = document.getElementById('search-box').value.toLowerCase().trim();
            const clickedMap = getClickedMap();
            list.innerHTML = '';

            let total = CONTACTS.length;
            let clickedCount = 0;

            // Hitung klik
            CONTACTS.forEach(c => {{
                if (clickedMap[c.tel]) clickedCount++;
            }});

            const filtered = CONTACTS.filter(c => {{
                const isClicked = !!clickedMap[c.tel];
                const matchesSearch = c.name.toLowerCase().includes(search) || c.tel.includes(search);
                const matchesFilter = !showOnlyUnclicked || !isClicked;
                return matchesSearch && matchesFilter;
            }});

            document.getElementById('total-cnt').innerText = total;
            document.getElementById('clicked-cnt').innerText = clickedCount;
            document.getElementById('left-cnt').innerText = Math.max(0, total - clickedCount);

            if (filtered.length === 0) {{
                list.innerHTML = `
                    <div class="empty-state">
                        <p style="font-size: 1.1rem; font-weight: 600; color: var(--text-title);">Data tidak ditemukan</p>
                        <p style="font-size: 0.9rem; margin-top: 0.25rem;">Coba ubah filter pencarian Anda.</p>
                    </div>
                `;
                return;
            }}

            filtered.forEach(c => {{
                const isClicked = !!clickedMap[c.tel];
                const card = document.createElement('div');
                card.className = 'contact-card' + (isClicked ? ' clicked' : '');
                
                card.innerHTML = `
                    <div class="contact-info">
                        <div class="contact-name">${{c.name}}</div>
                        <div class="contact-phone">${{c.tel}}</div>
                    </div>
                    <div class="action-area">
                        <span class="status-badge">${{isClicked ? 'SUDAH' : 'BELUM'}}</span>
                        <a href="${{c.url}}" target="_blank" class="wa-btn" onclick="trackClick('${{c.tel}}')">
                            💬 ${{isClicked ? 'Hubungi Lagi' : 'Hubungi'}}
                        </a>
                    </div>
                `;
                list.appendChild(card);
            }});
        }};

        const trackClick = (tel) => {{
            const map = getClickedMap();
            map[tel] = true;
            saveClickedMap(map);
            setTimeout(render, 300);
        }};

        document.getElementById('search-box').addEventListener('input', render);
        
        const filterBtn = document.getElementById('filter-btn');
        filterBtn.addEventListener('click', () => {{
            showOnlyUnclicked = !showOnlyUnclicked;
            if (showOnlyUnclicked) {{
                filterBtn.classList.add('active');
                filterBtn.innerHTML = '<span>👁️</span> Tampilkan Semua';
            }} else {{
                filterBtn.classList.remove('active');
                filterBtn.innerHTML = '<span>👁️</span> Tampilkan Belum Hubungi';
            }}
            render();
        }});

        const now = new Date();
        document.getElementById('date-label').innerText = 'Dashboard Follow-up — Dibuat pada ' + now.toLocaleDateString('id-ID', {{
            weekday: 'long', year: 'numeric', month: 'long', day: 'numeric'
        }});

        render();
    </script>
</body>
</html>
"""


async def cmd_walinkweb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not await require_member(update, context):
        return
    user_id = update.effective_user.id
    asyncio.create_task(adb.increment_usage(user_id))

    db.set_session(user_id, S0, {})
    from handlers.start import transition_to_handler
    await transition_to_handler(
        context.bot,
        user_id,
        update.effective_chat.id,
        "Kirim file <b>.xlsx</b>, <b>.csv</b>, <b>.txt</b>, atau <b>.vcf</b> sekarang.",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("BATAL & KEMBALI", callback_data="back_to_start", style="danger")]]),
        update=update
    )


async def handle_walinkweb_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    sess = db.get_session(user_id)
    if not sess or sess.get("state") != S0:
        return

    doc = update.message.document
    if not doc or not doc.file_name:
        await update.message.reply_text("Kirim file dokumen valid (.xlsx, .csv, .txt, .vcf).")
        return

    ext = os.path.splitext(doc.file_name)[1].lower()
    if ext not in (".xlsx", ".csv", ".txt", ".vcf"):
        await update.message.reply_text("Format tidak didukung. Kirim file .xlsx, .csv, .txt, atau .vcf.")
        return

    async with _get_lock(user_id):
        if user_id in _processing:
            return

    # Buat direktori temp untuk download
    user_dir = get_user_dir(user_id)
    work_dir = os.path.join(user_dir, f"walinkweb_{doc.file_id}")

    try:
        os.makedirs(work_dir, exist_ok=True)
        input_path = os.path.join(work_dir, f"input{ext}")
        file_obj = await context.bot.get_file(doc.file_id)
        await file_obj.download_to_drive(input_path)
        
        # Pindahkan ke state berikutnya: tanya pesan kustom
        data = {
            "input_path": input_path,
            "ext": ext,
            "file_name": doc.file_name,
            "work_dir": work_dir
        }
        db.set_session(user_id, S1, data)
        await update.message.reply_text(
            "Tulis pesan WhatsApp kustom untuk tautan follow-up?\n"
            "Ketik isi pesannya sekarang, atau ketik <b>-</b> (tanda minus) untuk mengosongkan tanpa pesan.",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("BATAL & KEMBALI", callback_data="back_to_start", style="danger")]])
        )
    except Exception as e:
        logger.error("Failed to download file in walinkweb: %s", e)
        shutil.rmtree(work_dir, ignore_errors=True)
        await update.message.reply_text("Gagal mengunduh berkas. Coba kirim ulang.")


async def handle_walinkweb_msg(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    sess = db.get_session(user_id)
    if not sess or sess.get("state") != S1:
        return

    msg_text = update.message.text.strip()
    # Jika input adalah "-", kosongkan pesan
    custom_msg = "" if msg_text == "-" else msg_text

    data = sess["data"]
    input_path = data["input_path"]
    ext = data["ext"]
    orig_name = data["file_name"]
    work_dir = data["work_dir"]

    async with _get_lock(user_id):
        if user_id in _processing:
            return
        _processing.add(user_id)
    db.clear_session(user_id)

    status_msg = await update.message.reply_text(
        "Memproses dan menyusun Dashboard Web interaktif...",
        parse_mode="HTML"
    )

    try:
        loop = asyncio.get_running_loop()

        def do_process():
            contacts = _extract_contacts_sync(input_path, ext)
            if not contacts:
                return 0, None, None

            # 1. Buat WhatsApp URL untuk setiap kontak
            encoded_msg = urllib.parse.quote(custom_msg) if custom_msg else ""
            
            formatted_contacts = []
            for c in contacts:
                url = f"https://wa.me/{c['tel']}?text={encoded_msg}" if encoded_msg else f"https://wa.me/{c['tel']}"
                formatted_contacts.append({
                    "name": c["name"],
                    "tel": c["tel"],
                    "url": url
                })

            # Buat page_id acak agar localStorage tidak bertabrakan dengan file lain
            import secrets
            page_id = secrets.token_hex(8)

            # Buat isi HTML
            contacts_json = json.dumps(formatted_contacts)
            html_content = HTML_TEMPLATE.format(
                filename=orig_name,
                contacts_json=contacts_json,
                page_id=page_id
            )

            html_buf = io.BytesIO(html_content.encode("utf-8"))

            # 2. Buat Excel workbook sebagai dukungan penuh / alternatif bagi pengguna iPhone
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "WhatsApp Links"

            # Style Header
            header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center")

            headers = ["No", "Nama Kontak", "Nomor HP", "Link WhatsApp"]
            ws.append(headers)

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            # Isi Baris
            row_align_center = Alignment(horizontal="center", vertical="center")
            row_align_left = Alignment(horizontal="left", vertical="center")
            
            for idx, c in enumerate(formatted_contacts, 1):
                ws.append([idx, c["name"], c["tel"], c["url"]])
                
                # Formatting link
                cell_link = ws.cell(row=idx+1, column=4)
                cell_link.hyperlink = c["url"]
                cell_link.font = Font(name="Arial", size=10, color="0000FF", underline="single")
                
                # Alignment
                ws.cell(row=idx+1, column=1).alignment = row_align_center
                ws.cell(row=idx+1, column=2).alignment = row_align_left
                ws.cell(row=idx+1, column=3).alignment = row_align_center
                ws.cell(row=idx+1, column=4).alignment = row_align_left

            # Atur Lebar Kolom otomatis
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

            excel_buf = io.BytesIO()
            wb.save(excel_buf)
            excel_buf.seek(0)

            return len(contacts), html_buf, excel_buf

        total_cnt, html_buf, excel_buf = await loop.run_in_executor(None, do_process)

        try:
            await status_msg.delete()
        except Exception:
            pass

        if total_cnt == 0:
            await update.message.reply_text("Tidak ada nomor HP valid yang ditemukan dalam berkas.")
            return

        base_name = os.path.splitext(orig_name)[0]
        out_name_html = f"DASHBOARD_WA_{base_name}.html"
        out_name_excel = f"BACKUP_WA_{base_name}.xlsx"
        
        html_buf.name = out_name_html
        excel_buf.name = out_name_excel

        from telegram import InputMediaDocument
        media_group = [
            InputMediaDocument(media=html_buf, filename=out_name_html),
            InputMediaDocument(media=excel_buf, filename=out_name_excel)
        ]

        await update.message.reply_media_group(
            media=media_group,
            read_timeout=120, write_timeout=120, connect_timeout=60
        )

        await update.message.reply_text(
            f"<b>Dashboard Web WA Links Selesai!</b>\n\n"
            f"Total kontak: <b>{total_cnt} nomor</b>\n"
            f"Pesan kustom: <i>{custom_msg if custom_msg else '[KOSONG]'}</i>\n\n"
            f"<b>Untuk Android / PC:</b>\n"
            f"Buka file <code>.html</code> di atas langsung di browser.\n\n"
            f"<b>Untuk iPhone (iOS) - 2 Cara Mudah:</b>\n"
            f"1. Buka file <code>.xlsx</code> (Excel) di atas yang otomatis didukung native oleh iPhone.\n"
            f"2. Atau, jika ingin tetap memakai file <code>.html</code>: ketuk file HTML -> ketuk tombol <b>Share/Bagikan</b> di kanan atas -> pilih <b>Safari/Chrome</b> (atau Simpan ke Files lalu buka via Safari).",
            parse_mode="HTML"
        )

        # Trigger debounced final keyboard
        old_timer = _button_timers.pop(user_id, None)
        if old_timer and not old_timer.done():
            old_timer.cancel()

        async def _send_buttons_debounced(uid, chat_id, bot):
            await asyncio.sleep(1.5)
            keyboard = InlineKeyboardMarkup([
                [
                    InlineKeyboardButton("PROSES FILE LAIN", callback_data="show_walinkweb_help", style="success"),
                    InlineKeyboardButton("KEMBALI KE MENU", callback_data="back_to_start", style="danger")
                ]
            ])
            from handlers.start import clear_welcome_messages
            clear_welcome_messages(uid)
            await bot.send_message(
                chat_id=chat_id,
                text="Proses selesai. Silakan unduh berkas dashboard HTML di atas.",
                reply_markup=keyboard
            )

        task = asyncio.create_task(_send_buttons_debounced(user_id, update.effective_chat.id, context.bot))
        _button_timers[user_id] = task

    except Exception as e:
        logger.error("Error in walinkweb handle processing: %s", e, exc_info=True)
        try:
            await update.message.reply_text("Terjadi kesalahan saat membuat dashboard HTML. Coba lagi.")
        except Exception:
            pass
    finally:
        _processing.discard(user_id)
        shutil.rmtree(work_dir, ignore_errors=True)


async def handle_show_walinkweb_help_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    user_id = query.from_user.id
    db.set_session(user_id, S0, {})

    try:
        await query.message.edit_text(
            text="Kirim file <b>.xlsx</b>, <b>.csv</b>, <b>.txt</b>, atau <b>.vcf</b> sekarang.",
            parse_mode="HTML"
        )
    except Exception:
        try:
            await query.message.delete()
        except Exception:
            pass
        await context.bot.send_message(
            chat_id=query.message.chat_id,
            text="Kirim file <b>.xlsx</b>, <b>.csv</b>, <b>.txt</b>, atau <b>.vcf</b> sekarang.",
            parse_mode="HTML",
            reply_markup=ReplyKeyboardRemove()
        )
